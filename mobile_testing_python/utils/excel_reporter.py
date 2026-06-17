import pandas as pd
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

def generate_professional_report(platform, test_results):
    reports_dir = '../reports'
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir)
        
    filename = f"{reports_dir}/{platform}_E2E_Analysis.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "E2E Analysis"

    # Define Header
    headers = ["#", "Category", "Test Case", "Status", "Error Detail", "Timestamp"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="2D3949", end_color="2D3949", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    category_colors = {
        'Functional Testing': 'E8F5E9',
        'UI/UX Testing': 'E3F2FD',
        'Compatibility Testing': 'F1F8E9',
        'Performance Testing': 'FCE4EC',
        'Security Testing': 'F3E5F5',
        'API Testing': 'E0F2F1',
        'Database Testing': 'EFEBE9',
        'Accessibility Testing': 'FAFAFA',
        'Mobile-Specific Testing': 'ECEFF1',
        'Regression Testing': 'FFF9C4',
        'End-to-End Testing': 'FFEBEE'
    }

    for idx, result in enumerate(test_results, 1):
        row = [
            idx,
            result.get('category', 'N/A'),
            result.get('title', 'N/A'),
            result.get('status', 'N/A'),
            result.get('error', ''),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ]
        ws.append(row)
        
        # Style row
        row_idx = ws.max_row
        bg_color = category_colors.get(result.get('category'), 'FFFFFF')
        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
        for cell in ws[row_idx]:
            cell.fill = fill
            if cell.column == 4: # Status
                if result['status'] == 'PASS':
                    cell.font = Font(color="2E7D32", bold=True)
                else:
                    cell.font = Font(color="C62828", bold=True)

    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['E'].width = 40
    
    wb.save(filename)
    print(f"✅ Professional Mobile Report Generated: {filename}")
